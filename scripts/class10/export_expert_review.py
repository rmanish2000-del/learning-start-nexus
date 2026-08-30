#!/usr/bin/env python3
"""Export the 326 Class 10 (2026-27) draft items into subject-expert review packages.

  python3 scripts/class10/export_expert_review.py [outdir]

Read-only with respect to the database. No learner, parent or payment data is
touched: the source is EDUOS_CLASS_10_FINAL_QUESTION_REGISTER.json only.
"""
import csv
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "review-bundles" / "class10-2026-27-expert"

SYSTEM_COLS = [
    ("Question ID", "externalRef", 26),
    ("Subject", "subject", 13),
    ("Unit", "unitTitle", 24),
    ("Chapter", "chapterTitle", 24),
    ("Official requirement", None, 30),
    ("Outcome", None, 40),
    ("Atom/concept", "topicTitle", 26),
    ("Question", "prompt", 60),
    ("Options", None, 34),
    ("Correct answer", "correctAnswer", 24),
    ("Explanation", "explanation", 60),
    ("Difficulty", "difficulty", 10),
    ("Source reference", "officialSourceReference", 30),
    ("Current status", None, 22),
]
EXPERT_COLS = [
    ("Expert decision", 16),
    ("Expert correction", 40),
    ("Expert comments", 40),
    ("Reviewer name", 20),
    ("Qualification", 22),
    ("Review date", 14),
]

HEADER_FILL = PatternFill("solid", start_color="1F3864")
EXPERT_FILL = PatternFill("solid", start_color="C6E0B4")
LOCKED_FILL = PatternFill("solid", start_color="F2F2F2")
FONT = "Arial"


def load():
    data = json.loads((ROOT / "EDUOS_CLASS_10_FINAL_QUESTION_REGISTER.json").read_text())
    return data["items"]


def cell_value(item, header, key):
    if key:
        v = item.get(key)
        return "" if v is None else v
    if header == "Official requirement":
        return ", ".join(item.get("officialRequirementIds") or [])
    if header == "Outcome":
        return f"{item.get('outcomeCode','')} — {item.get('outcomeTitle','')}".strip(" —")
    if header == "Options":
        opts = item.get("options")
        return "\n".join(f"{chr(65+i)}. {o}" for i, o in enumerate(opts)) if opts else "(constructed response)"
    if header == "Current status":
        return f"{item.get('status')} / {item.get('verificationState')} / {item.get('reviewStatus')}"
    return ""


def style_sheet(ws, rows, headers, expert_start):
    for idx, (header, width) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=idx, value=header)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{rows + 1}"
    for r in range(2, rows + 2):
        ws.row_dimensions[r].height = 46
        for idx in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=idx)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            editable = idx >= expert_start
            c.protection = c.protection.copy(locked=not editable)
            if not editable:
                c.fill = LOCKED_FILL
            else:
                c.fill = EXPERT_FILL


def build_subject_workbook(subject, items, path):
    headers = [(h, w) for h, _, w in SYSTEM_COLS] + EXPERT_COLS
    expert_start = len(SYSTEM_COLS) + 1
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    for item in items:
        ws.append(
            [cell_value(item, h, k) for h, k, _ in SYSTEM_COLS]
            + ["", "", "", "", "", ""]
        )
    style_sheet(ws, len(items), headers, expert_start)

    dv = DataValidation(type="list", formula1='"APPROVE,REVISE,REJECT"', allow_blank=True, showErrorMessage=True)
    dv.error = "Choose APPROVE, REVISE or REJECT."
    ws.add_data_validation(dv)
    col = get_column_letter(expert_start)
    dv.add(f"{col}2:{col}{len(items) + 1}")
    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.password = None
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatCells = False

    # Summary sheet — counts computed with Excel formulas over the Review sheet.
    s = wb.create_sheet("Summary")
    s["A1"] = f"{subject} — Class 10 CBSE 2026-27 draft review"
    s["A1"].font = Font(name=FONT, bold=True, size=13)
    s["A3"] = "Total questions"
    s["B3"] = f"=COUNTA(Review!A2:A{len(items) + 1})"
    s["A4"] = "Approved"
    s["B4"] = f'=COUNTIF(Review!O2:O{len(items) + 1},"APPROVE")'
    s["A5"] = "Revise"
    s["B5"] = f'=COUNTIF(Review!O2:O{len(items) + 1},"REVISE")'
    s["A6"] = "Rejected"
    s["B6"] = f'=COUNTIF(Review!O2:O{len(items) + 1},"REJECT")'
    s["A7"] = "Undecided"
    s["B7"] = "=B3-B4-B5-B6"
    s["A9"] = "Unit"
    s["B9"] = "Questions"
    s["C9"] = "Approved"
    s["D9"] = "Revise"
    s["E9"] = "Rejected"
    for c in ("A9", "B9", "C9", "D9", "E9"):
        s[c].font = Font(name=FONT, bold=True)
    units = sorted({i["unitTitle"] for i in items})
    row = 10
    for u in units:
        s.cell(row=row, column=1, value=u)
        s.cell(row=row, column=2, value=f'=COUNTIF(Review!C2:C{len(items) + 1},A{row})')
        s.cell(row=row, column=3, value=f'=COUNTIFS(Review!C2:C{len(items) + 1},A{row},Review!O2:O{len(items) + 1},"APPROVE")')
        s.cell(row=row, column=4, value=f'=COUNTIFS(Review!C2:C{len(items) + 1},A{row},Review!O2:O{len(items) + 1},"REVISE")')
        s.cell(row=row, column=5, value=f'=COUNTIFS(Review!C2:C{len(items) + 1},A{row},Review!O2:O{len(items) + 1},"REJECT")')
        row += 1
    for col_i, w in ((1, 40), (2, 12), (3, 12), (4, 12), (5, 12)):
        s.column_dimensions[get_column_letter(col_i)].width = w
    for r in s.iter_rows(min_row=1, max_row=row):
        for c in r:
            if c.value is not None and not c.font.bold:
                c.font = Font(name=FONT, size=10)
    s.protection.sheet = True

    wb.save(path)


def sha256(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    items = load()
    maths = [i for i in items if i["subject"] == "Mathematics"]
    science = [i for i in items if i["subject"] == "Science"]
    ids = [i["externalRef"] for i in items]
    assert len(ids) == len(set(ids)), "duplicate question ids"
    assert len(maths) + len(science) == len(items) == 326, "count mismatch"

    m_path = OUT / "EDUOS_CLASS10_MATHEMATICS_EXPERT_REVIEW.xlsx"
    s_path = OUT / "EDUOS_CLASS10_SCIENCE_EXPERT_REVIEW.xlsx"
    build_subject_workbook("Mathematics", maths, m_path)
    build_subject_workbook("Science", science, s_path)

    c_path = OUT / "EDUOS_CLASS10_EXPERT_DECISIONS_IMPORT_TEMPLATE.csv"
    with c_path.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([
            "question_id", "subject", "unit", "expert_decision", "expert_correction",
            "expert_comments", "reviewer_name", "qualification", "review_date",
        ])
        for i in items:
            w.writerow([i["externalRef"], i["subject"], i["unitTitle"], "", "", "", "", "", ""])

    summary = {
        "mathematics": len(maths),
        "science": len(science),
        "total": len(items),
        "duplicates": 0,
        "missing": 0,
        "files": {p.name: sha256(p) for p in (m_path, s_path, c_path)},
    }
    print(json.dumps(summary, indent=2))
    return summary


if __name__ == "__main__":
    main()
